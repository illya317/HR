#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
车间报表 Excel 转 JSON 校验脚本
校验范围：文件级、财务报表 JSON、工分明细 JSON、关联 JSON
输出：JSON 报告 + Markdown 报告 + 可选 CSV
"""

import argparse
import csv
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# 常量与映射
# ---------------------------------------------------------------------------
NORMALIZATION_MAP = {
    "盐酸左氧氟沙星胶囊": "盐酸氧氟沙星胶囊",
    "甘草酸二胺胶囊": "甘草酸二铵胶囊",
}

ISSUE_SEVERITY_ORDER = {"ERROR": 0, "WARN": 1, "INFO": 2}


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------
def clean(val):
    if pd.isna(val):
        return None
    if isinstance(val, (int, float)):
        if pd.isna(val):
            return None
        return val
    s = str(val).strip()
    if s == "nan" or s == "":
        return None
    return s


def try_num(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    try:
        return float(val)
    except (ValueError, TypeError):
        return val


def normalize_name(name):
    if not name:
        return ""
    name = name.strip()
    for k, v in NORMALIZATION_MAP.items():
        name = name.replace(k, v)
    return name


def _parse_segment(start_num_str, seg):
    """Parse a single segment like '20251201-1205' into list of ints."""
    numbers = [int(start_num_str)]
    range_match = re.search(r"[-–—]\s*(\d+)", seg)
    if range_match:
        end_str = range_match.group(1)
        prefix_len = len(start_num_str) - len(end_str)
        if prefix_len > 0:
            end_num = int(start_num_str[:prefix_len] + end_str)
        else:
            end_num = int(end_str)
        start_num = int(start_num_str)
        if end_num > start_num:
            if str(start_num)[:6] == str(end_num)[:6]:
                for n in range(start_num + 1, end_num + 1):
                    numbers.append(n)
            else:
                numbers.append(end_num)
    return numbers


def _parse_standalone_segment(seg, reference_num):
    """Parse a standalone segment like '20250224-26', '18', or '0313-0314'."""
    # Full number with optional range: 20250224-26
    m = re.match(r"(\d{8})(?:[-–—](\d+))?", seg)
    if m:
        start = int(m.group(1))
        end_str = m.group(2)
        if end_str:
            prefix_len = 8 - len(end_str)
            if prefix_len > 0:
                end = int(str(start)[:prefix_len] + end_str)
            else:
                end = int(end_str)
            if end >= start:
                return list(range(start, end + 1))
        return [start]

    # Short number with optional range using reference prefix: 0313-0314 → 20260313-20260314
    m = re.match(r"(\d+)(?:[-–—](\d+))?", seg)
    if m:
        short_start = m.group(1)
        short_end = m.group(2)
        prefix_len = len(reference_num) - len(short_start)
        if prefix_len > 0:
            start = int(reference_num[:prefix_len] + short_start)
        else:
            start = int(short_start)
        if short_end:
            end_prefix_len = len(reference_num) - len(short_end)
            if end_prefix_len > 0:
                end = int(reference_num[:end_prefix_len] + short_end)
            else:
                end = int(short_end)
            if end >= start:
                return list(range(start, end + 1))
        return [start]
    return []


def parse_batch_numbers(batch_str):
    """Extract batch numbers from work detail string. Supports / and 、 separators."""
    if not batch_str:
        return []
    m = re.search(r"批号[：:]?\s*(\d+)", batch_str)
    if not m:
        m = re.search(r"\b(\d{8,})\b", batch_str)
        if not m:
            return []
    full_num = m.group(1)
    if len(full_num) < 8:
        return []

    suffix = batch_str[m.end():]
    numbers = []

    # Split by / or 、 to handle segments like "20250212/20250224-26"
    segments = re.split(r"[/、]", suffix)
    first_seg = segments[0] if segments else ""

    # First segment is attached to full_num
    numbers.extend(_parse_segment(full_num, first_seg))

    # Subsequent segments are standalone
    for seg in segments[1:]:
        seg = seg.strip()
        if seg:
            numbers.extend(_parse_standalone_segment(seg, full_num))

    return sorted(list(set(numbers)))


def is_number(v):
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(v)
        return True
    except (ValueError, TypeError):
        return False


# ---------------------------------------------------------------------------
# Excel 独立抽取
# ---------------------------------------------------------------------------
class ExcelExtractor:
    def __init__(self, excel_path):
        self.path = Path(excel_path)
        self.wb = openpyxl.load_workbook(self.path, data_only=True)
        self.issues = []

    def _add_issue(self, severity, type_, message, **kwargs):
        issue = {
            "severity": severity,
            "type": type_,
            "message": message,
        }
        issue.update(kwargs)
        self.issues.append(issue)

    def extract_finance(self):
        """动态查找财务报表 sheet 并抽取数据。"""
        sheet_name = None
        for sn in self.wb.sheetnames:
            if re.match(r"\d{4}年财务报表", sn):
                sheet_name = sn
                break
        if sheet_name is None:
            self._add_issue("ERROR", "sheet_missing", "缺少财务报表 sheet（如 2025年财务报表）")
            return None

        ws = self.wb[sheet_name]
        months = []
        current_month = None

        for row in ws.iter_rows(min_row=1, values_only=False):
            first_cell = row[0]
            first_val = clean(first_cell.value)

            if first_val and isinstance(first_val, str) and re.match(r"\d{4}年\d{2}月片剂车间生产品种入库报表", first_val):
                m = re.search(r"(\d{2})月", first_val)
                month_num = int(m.group(1)) if m else None
                current_month = {
                    "月份": month_num,
                    "报表标题": first_val,
                    "品种记录": [],
                    "合计": None,
                }
                continue

            if current_month is None:
                continue

            if first_val == "合   计" or (isinstance(first_val, str) and "合计" in first_val):
                current_month["合计"] = {
                    "投料量_万粒": try_num(clean(row[4].value)),
                    "工分": try_num(clean(row[5].value)),
                    "生产数量": clean(row[6].value),
                    "折合_盒瓶": try_num(clean(row[7].value)),
                    "折合_万粒": try_num(clean(row[8].value)),
                }
                months.append(current_month)
                current_month = None
                continue

            # Skip non-data rows
            if first_val and isinstance(first_val, str) and ("制表" in first_val or "复核" in first_val or "制 表" in first_val):
                continue

            seq = clean(row[0].value)
            name = clean(row[1].value)
            spec = clean(row[2].value)
            batch = clean(row[3].value)
            feed = try_num(clean(row[4].value))
            points = try_num(clean(row[5].value))
            qty = clean(row[6].value)
            box = try_num(clean(row[7].value))
            wanli = try_num(clean(row[8].value))
            remark = clean(row[9].value)

            if seq is not None and (isinstance(seq, int) or (isinstance(seq, str) and seq.isdigit())):
                current_product = {
                    "序号": int(seq) if isinstance(seq, str) else seq,
                    "品种": name,
                    "规格": spec,
                    "批次记录": [],
                }
                current_month["品种记录"].append(current_product)

            if current_month["品种记录"]:
                prod = current_month["品种记录"][-1]
                batch_rec = {
                    "批号": try_num(batch),
                    "投料量_万粒": feed,
                    "工分": points,
                    "生产数量": qty,
                    "折合_盒瓶": box,
                    "折合_万粒": wanli,
                    "备注": remark,
                }
                batch_rec = {k: v for k, v in batch_rec.items() if v is not None}
                if batch_rec or (first_val is not None):
                    prod["批次记录"].append(batch_rec)

        return months

    def extract_work_details(self):
        """从 'X月工分明细' sheets 抽取数据。"""
        result = {}
        for sheet_name in self.wb.sheetnames:
            m = re.search(r"(\d{1,2})月工分明细", sheet_name)
            if not m:
                continue
            month_num = int(m.group(1))
            ws = self.wb[sheet_name]
            products = []
            current_product = None
            current_entries = []
            current_position = None

            for row in ws.iter_rows(min_row=1, values_only=False):
                vals = [clean(c.value) for c in row]
                c0 = vals[0]

                if c0 and isinstance(c0, str) and "固体制剂车间入库品种工分明细" in c0:
                    if current_product is not None:
                        products.append(current_product)
                    current_product = None
                    current_entries = []
                    current_position = None
                    continue

                if current_product is None and c0 and isinstance(c0, str) and (c0.startswith("品名：") or c0.startswith("品名:")):
                    name = re.sub(r"^品名[：:]", "", c0).strip()
                    batch_info = vals[3] if len(vals) > 3 and vals[3] and isinstance(vals[3], str) and "批号" in vals[3] else None
                    current_product = {
                        "品名": name,
                        "批号信息": batch_info,
                        "人员工分明细": [],
                    }
                    continue

                if current_product is None:
                    continue

                # Header row
                if c0 and isinstance(c0, str) and "序" in c0 and "号" in c0:
                    continue

                # Total row
                total_val = vals[10] if len(vals) > 10 else None
                if c0 is None and vals[1] is None and vals[2] is None and total_val is not None and is_number(total_val):
                    try:
                        float(total_val)
                        # Save current product
                        if current_product is not None:
                            products.append(current_product)
                            current_product = None
                        continue
                    except (ValueError, TypeError):
                        pass

                # Next product block (without title row in between)
                if c0 and isinstance(c0, str) and (c0.startswith("品名：") or c0.startswith("品名:")):
                    if current_product is not None:
                        products.append(current_product)
                    name = re.sub(r"^品名[：:]", "", c0).strip()
                    batch_info = vals[3] if len(vals) > 3 and vals[3] and isinstance(vals[3], str) and "批号" in vals[3] else None
                    current_product = {
                        "品名": name,
                        "批号信息": batch_info,
                        "人员工分明细": [],
                    }
                    current_position = None
                    continue

                name = vals[2] if len(vals) > 2 else None
                if name is None:
                    continue

                if c0 is not None and isinstance(c0, str) and c0.strip() != "":
                    current_position = c0.strip()

                position_unit_price = try_num(vals[1]) if len(vals) > 1 else None
                daily_scores = []
                for col_idx in range(3, 9):
                    v = vals[col_idx] if col_idx < len(vals) else None
                    if v is not None and is_number(v):
                        daily_scores.append(try_num(v))
                remark = vals[9] if len(vals) > 9 else None
                total = try_num(vals[10]) if len(vals) > 10 else None

                entry = {
                    "岗位类别": current_position,
                    "姓名": name,
                    "岗位工分单价": position_unit_price,
                    "每日工分": daily_scores,
                    "备注": remark,
                    "合计": total,
                }
                entry = {k: v for k, v in entry.items() if v is not None}
                if entry:
                    current_product["人员工分明细"].append(entry)

            if current_product is not None:
                products.append(current_product)

            result[f"{month_num}月"] = products

        return result


# ---------------------------------------------------------------------------
# 校验引擎
# ---------------------------------------------------------------------------
class Validator:
    def __init__(self, excel_path, finance_json, work_json, linked_json):
        self.excel_path = Path(excel_path)
        self.finance_json = Path(finance_json)
        self.work_json = Path(work_json)
        self.linked_json = Path(linked_json)
        self.issues = []
        self.summary = {}
        self.month_summary = []

    def _add_issue(self, severity, type_, message, **kwargs):
        issue = {
            "severity": severity,
            "type": type_,
            "message": message,
        }
        issue.update(kwargs)
        self.issues.append(issue)

    def _load_json(self, path, name):
        if not path.exists():
            self._add_issue("ERROR", "file_missing", f"{name} 不存在: {path}")
            return None
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            self._add_issue("ERROR", "json_parse_error", f"{name} 解析失败: {e}")
            return None

    def run(self):
        # File-level
        fin_data = self._load_json(self.finance_json, "财务报表 JSON")
        work_data = self._load_json(self.work_json, "工分明细 JSON")
        linked_data = self._load_json(self.linked_json, "关联 JSON")

        if not all([fin_data, work_data, linked_data]):
            return

        # Excel independent extraction
        extractor = ExcelExtractor(self.excel_path)
        excel_finance = extractor.extract_finance()
        excel_work = extractor.extract_work_details()

        if extractor.issues:
            self.issues.extend(extractor.issues)

        # Determine expected months from Excel data
        expected_months = set()
        if excel_finance:
            expected_months.update(m["月份"] for m in excel_finance if m.get("月份"))
        if excel_work:
            expected_months.update(int(k.replace("月", "")) for k in excel_work.keys() if k.replace("月", "").isdigit())
        if not expected_months:
            # fallback: use JSON months
            expected_months = set(m["月份"] for m in fin_data if "月份" in m)
            expected_months.update(int(k.replace("月", "")) for k in work_data.keys() if k.replace("月", "").isdigit())

        self._expected_months = sorted(expected_months)

        # Validate finance JSON
        self._validate_finance_json(fin_data, excel_finance)

        # Validate work JSON
        self._validate_work_json(work_data, excel_work)

        # Validate linked JSON
        self._validate_linked_json(linked_data, fin_data, work_data)

        # Month-level summary
        self._build_month_summary(fin_data, work_data, linked_data)

        # Overall status
        errors = [i for i in self.issues if i["severity"] == "ERROR"]
        warns = [i for i in self.issues if i["severity"] == "WARN"]
        if errors:
            status = "FAIL"
        elif warns:
            status = "REVIEW"
        else:
            status = "PASS"

        base_summary = {
            "status": status,
            "total_issues": len(self.issues),
            "errors": len(errors),
            "warns": len(warns),
            "infos": len([i for i in self.issues if i["severity"] == "INFO"]),
        }
        # Merge with fields already set by _validate_linked_json
        self.summary = {**base_summary, **self.summary}

    def _validate_finance_json(self, fin_data, excel_finance):
        # Month check
        months_found = {m["月份"] for m in fin_data if "月份" in m}
        for m in self._expected_months:
            if m not in months_found:
                self._add_issue("ERROR", "month_missing", f"财务报表 JSON 缺少 {m} 月")

        # Compare with Excel
        if excel_finance:
            excel_months = {m["月份"] for m in excel_finance}
            for m in self._expected_months:
                if m not in excel_months:
                    self._add_issue("ERROR", "excel_month_missing", f"Excel 财务报表缺少 {m} 月")

            for m in fin_data:
                month = m["月份"]
                excel_m = next((x for x in excel_finance if x["月份"] == month), None)
                if not excel_m:
                    continue

                json_batches = sum(len(p["批次记录"]) for p in m.get("品种记录", []))
                excel_batches = sum(len(p["批次记录"]) for p in excel_m.get("品种记录", []))
                if json_batches != excel_batches:
                    self._add_issue(
                        "ERROR",
                        "batch_count_mismatch",
                        f"{month}月 财务批次数不一致: JSON={json_batches}, Excel={excel_batches}",
                        month=month,
                        json_batches=json_batches,
                        excel_batches=excel_batches,
                    )

                # Totals
                j_total = m.get("合计", {})
                e_total = excel_m.get("合计", {})
                for key in ["投料量_万粒", "工分", "折合_盒瓶", "折合_万粒"]:
                    jv = j_total.get(key)
                    ev = e_total.get(key)
                    if jv != ev and (jv is not None or ev is not None):
                        self._add_issue(
                            "WARN",
                            "total_mismatch",
                            f"{month}月 合计 {key} 不一致: JSON={jv}, Excel={ev}",
                            month=month,
                            key=key,
                            json_value=jv,
                            excel_value=ev,
                        )

        # Record-level checks
        for m in fin_data:
            month = m.get("月份")
            for prod in m.get("品种记录", []):
                if not prod.get("品种"):
                    self._add_issue("ERROR", "empty_product_name", f"{month}月 存在空品种名", month=month)
                for batch in prod.get("批次记录", []):
                    if batch.get("批号") is None:
                        self._add_issue(
                            "WARN",
                            "empty_batch_no",
                            f"{month}月 {prod.get('品种')} 存在空批号批次",
                            month=month,
                            product=prod.get("品种"),
                        )
                    if batch.get("工分") is None:
                        self._add_issue(
                            "WARN",
                            "empty_batch_points",
                            f"{month}月 {prod.get('品种')} 批号{batch.get('批号')} 工分为空",
                            month=month,
                            product=prod.get("品种"),
                            batch_no=batch.get("批号"),
                        )

    def _validate_work_json(self, work_data, excel_work):
        for m in self._expected_months:
            month_key = f"{m}月"
            if month_key not in work_data:
                self._add_issue("ERROR", "work_month_missing", f"工分明细 JSON 缺少 {month_key}")

        if excel_work:
            for month_key, excel_products in excel_work.items():
                json_products = work_data.get(month_key, [])
                if len(json_products) != len(excel_products):
                    self._add_issue(
                        "ERROR",
                        "work_block_count_mismatch",
                        f"{month_key} 工分明细块数不一致: JSON={len(json_products)}, Excel={len(excel_products)}",
                        month=month_key,
                        json_count=len(json_products),
                        excel_count=len(excel_products),
                    )

        for month_key, products in work_data.items():
            month_num = int(month_key.replace("月", ""))
            for block in products:
                if not block.get("品名"):
                    self._add_issue(
                        "WARN",
                        "work_empty_product",
                        f"{month_key} 存在空品名工分块",
                        month=month_num,
                    )
                people = block.get("人员工分明细", [])
                block_total_from_excel = None
                for person in people:
                    if not person.get("姓名"):
                        self._add_issue(
                            "ERROR",
                            "work_empty_name",
                            f"{month_key} {block.get('品名')} 存在空姓名",
                            month=month_num,
                            product=block.get("品名"),
                        )
                    daily = person.get("每日工分", [])
                    person_total = person.get("合计")
                    if daily and person_total is not None:
                        s = sum(daily)
                        if abs(s - person_total) > 0.01:
                            self._add_issue(
                                "WARN",
                                "daily_sum_mismatch",
                                f"{month_key} {block.get('品名')} {person.get('姓名')} 每日工分求和 {s} != 合计 {person_total}",
                                month=month_num,
                                product=block.get("品名"),
                                person=person.get("姓名"),
                                daily_sum=s,
                                person_total=person_total,
                            )

    def _validate_linked_json(self, linked_data, fin_data, work_data):
        # Count total batches in finance vs linked
        fin_batches = sum(
            len(p["批次记录"]) for m in fin_data for p in m.get("品种记录", [])
        )
        linked_batches = sum(
            len(p["批次记录"]) for m in linked_data for p in m.get("品种记录", [])
        )
        if fin_batches != linked_batches:
            self._add_issue(
                "ERROR",
                "linked_batch_count_mismatch",
                f"关联 JSON 批次数与财务报表不一致: finance={fin_batches}, linked={linked_batches}",
            )

        # Check unmatched
        total_unmatched = 0
        for m in linked_data:
            um = m.get("未匹配工分明细", [])
            if um:
                total_unmatched += len(um)
                self._add_issue(
                    "ERROR" if len(um) > 0 else "INFO",
                    "unmatched_work_details",
                    f"{m['月份']}月 有 {len(um)} 条未匹配工分明细",
                    month=m["月份"],
                    count=len(um),
                )

        # Count work items in linked vs work json
        work_total_items = sum(len(v) for v in work_data.values())
        linked_people = 0
        mismatch_count = 0

        # Cross-batch fingerprint tracking for multi_batch_detail
        detail_locations = {}  # fingerprint -> list of locations

        for m in linked_data:
            month = m["月份"]
            for p in m.get("品种记录", []):
                for b in p.get("批次记录", []):
                    wds = b.get("工分明细", [])
                    linked_people += sum(len(wd["人员工分明细"]) for wd in wds)

                    for wd in wds:
                        fp = (wd.get("品名"), wd.get("批号信息"))
                        loc = {
                            "month": month,
                            "product": p.get("品种"),
                            "batch_no": b.get("批号"),
                        }
                        detail_locations.setdefault(fp, []).append(loc)

                    # Check finance points vs people points
                    fin_points = b.get("工分")
                    people_total = sum(
                        sum(person.get("合计", 0) for person in wd.get("人员工分明细", []))
                        for wd in wds
                    )
                    if fin_points is not None and wds and people_total != 0:
                        if abs(fin_points - people_total) > 0.01:
                            mismatch_count += 1
                            self._add_issue(
                                "WARN",
                                "batch_work_points_mismatch",
                                f"{month}月 {p.get('品种')} 批号{b.get('批号')} 财务工分 {fin_points} != 人员合计 {people_total}",
                                month=month,
                                product=p.get("品种"),
                                batch_no=b.get("批号"),
                                finance_points=fin_points,
                                people_points=people_total,
                                diff=round(fin_points - people_total, 2),
                            )

        # Report multi_batch_detail issues based on cross-batch fingerprints
        multi_batch_count = 0
        for fp, locations in detail_locations.items():
            if len(locations) > 1:
                multi_batch_count += 1
                locs_str = ", ".join(
                    f"{loc['month']}月批号{loc['batch_no']}" for loc in locations
                )
                self._add_issue(
                    "WARN",
                    "multi_batch_detail",
                    f"工分明细 '{fp[0]} | {fp[1]}' 被挂到 {len(locations)} 个财务批次: {locs_str}",
                    product=fp[0],
                    batch_info=fp[1],
                    locations=locations,
                )

        linked_unique_items = len(detail_locations)
        if linked_unique_items != work_total_items:
            diff = work_total_items - linked_unique_items
            # Check if the discrepancy is due to cross-month duplicates
            # (same work detail fingerprint appearing in multiple months)
            cross_month_dups = []
            work_fp_months = {}
            for month_key, products in work_data.items():
                for p in products:
                    fp = (p.get("品名"), p.get("批号信息"))
                    work_fp_months.setdefault(fp, []).append(month_key)
            for fp, months in work_fp_months.items():
                if len(months) > 1:
                    cross_month_dups.append({"fp": fp, "months": months})

            if diff > 0 and len(cross_month_dups) == diff:
                # The entire discrepancy is explained by cross-month duplicates
                for dup in cross_month_dups:
                    self._add_issue(
                        "WARN",
                        "cross_month_work_detail_duplicate",
                        f"工分明细 '{dup['fp'][0]} | {dup['fp'][1]}' 在 {', '.join(dup['months'])} 重复出现",
                        product=dup["fp"][0],
                        batch_info=dup["fp"][1],
                        months=dup["months"],
                    )
            else:
                severity = "ERROR" if diff > 0 else "WARN"
                self._add_issue(
                    severity,
                    "linked_work_item_count_mismatch",
                    f"关联 JSON 唯一工分明细数 {linked_unique_items} != 工分明细 JSON {work_total_items} (差异 {diff})",
                    linked_unique_items=linked_unique_items,
                    work_total_items=work_total_items,
                    diff=diff,
                )

        self.summary["multi_batch_detail_count"] = multi_batch_count
        self.summary["batch_points_mismatch_count"] = mismatch_count
        self.summary["total_people_matched"] = linked_people

    def _build_month_summary(self, fin_data, work_data, linked_data):
        for m in self._expected_months:
            fin_m = next((x for x in fin_data if x.get("月份") == m), None)
            linked_m = next((x for x in linked_data if x.get("月份") == m), None)
            wkey = f"{m}月"
            work_products = work_data.get(wkey, [])

            fin_batches = sum(len(p["批次记录"]) for p in fin_m.get("品种记录", [])) if fin_m else 0
            fin_products = len(fin_m.get("品种记录", [])) if fin_m else 0
            fin_total_points = (fin_m.get("合计") or {}).get("工分") if fin_m else None
            if fin_total_points is None and fin_m:
                # Fallback: sum from batch records when total row is empty
                fin_total_points = round(sum(
                    (b.get("工分") or 0)
                    for p in fin_m.get("品种记录", [])
                    for b in p.get("批次记录", [])
                ), 2)

            work_blocks = len(work_products)
            work_people = sum(len(p.get("人员工分明细", [])) for p in work_products)

            linked_batches = sum(len(p["批次记录"]) for p in linked_m.get("品种记录", [])) if linked_m else 0
            linked_with_work = 0
            linked_people = 0
            unmatched = len(linked_m.get("未匹配工分明细", [])) if linked_m else 0

            if linked_m:
                for p in linked_m.get("品种记录", []):
                    for b in p.get("批次记录", []):
                        if "工分明细" in b:
                            linked_with_work += 1
                            linked_people += sum(len(wd["人员工分明细"]) for wd in b["工分明细"])

            self.month_summary.append({
                "月份": m,
                "财务品种数": fin_products,
                "财务批次数": fin_batches,
                "财务工分合计": fin_total_points,
                "工分明细块数": work_blocks,
                "工分明细人次": work_people,
                "关联批次数": linked_batches,
                "已挂载工分明细批次数": linked_with_work,
                "已匹配人员人次": linked_people,
                "未匹配工分明细数": unmatched,
            })

    def generate_json_report(self, out_path):
        report = {
            "status": self.summary.get("status", "FAIL"),
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "inputs": {
                "excel": str(self.excel_path),
                "finance_json": str(self.finance_json),
                "work_json": str(self.work_json),
                "linked_json": str(self.linked_json),
            },
            "summary": self.summary,
            "month_summary": self.month_summary,
            "issues": self.issues,
        }
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        return report

    def generate_md_report(self, out_path, report):
        lines = []
        lines.append("# 车间报表 JSON 转换校验报告")
        lines.append("")
        lines.append(f"**生成时间**: {report['generated_at']}")
        lines.append(f"**状态**: {report['status']}")
        lines.append("")
        lines.append("## 输入文件")
        lines.append("")
        for k, v in report["inputs"].items():
            lines.append(f"- `{k}`: `{v}`")
        lines.append("")
        lines.append("## 汇总")
        lines.append("")
        lines.append(f"| 指标 | 数值 |")
        lines.append(f"|---|---:|")
        for k, v in report["summary"].items():
            lines.append(f"| {k} | {v} |")
        lines.append("")
        lines.append("## 月度明细")
        lines.append("")
        if report["month_summary"]:
            headers = list(report["month_summary"][0].keys())
            lines.append("| " + " | ".join(headers) + " |")
            lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
            for row in report["month_summary"]:
                vals = [str(row.get(h, "")) for h in headers]
                lines.append("| " + " | ".join(vals) + " |")
        lines.append("")

        # Issues by severity
        for severity in ["ERROR", "WARN", "INFO"]:
            sev_issues = [i for i in report["issues"] if i["severity"] == severity]
            if not sev_issues:
                continue
            lines.append(f"## {severity} 列表 ({len(sev_issues)})")
            lines.append("")
            for idx, issue in enumerate(sev_issues, 1):
                lines.append(f"### {idx}. {issue['type']}")
                lines.append(f"- **消息**: {issue['message']}")
                for k, v in issue.items():
                    if k not in ("severity", "type", "message"):
                        lines.append(f"- **{k}**: {v}")
                lines.append("")

        with open(out_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

    def generate_csv_detail(self, out_path):
        fieldnames = ["severity", "type", "message", "month", "product", "batch_no", "finance_points", "people_points", "diff"]
        with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            for issue in self.issues:
                writer.writerow(issue)


# ---------------------------------------------------------------------------
# 主入口
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="车间报表 Excel 转 JSON 校验")
    parser.add_argument("--excel", required=True, help="Excel 源文件路径")
    parser.add_argument("--finance-json", required=True, help="财务报表 JSON")
    parser.add_argument("--work-json", required=True, help="工分明细 JSON")
    parser.add_argument("--linked-json", required=True, help="关联 JSON")
    parser.add_argument("--out-prefix", required=True, help="输出文件名前缀")
    parser.add_argument("--csv", action="store_true", help="同时输出 CSV")
    args = parser.parse_args()

    validator = Validator(args.excel, args.finance_json, args.work_json, args.linked_json)
    validator.run()

    base = Path(args.out_prefix)
    json_path = base.with_suffix(".json")
    md_path = base.with_suffix(".md")

    report = validator.generate_json_report(json_path)
    validator.generate_md_report(md_path, report)

    if args.csv:
        csv_path = base.with_suffix(".csv")
        validator.generate_csv_detail(csv_path)
        print(f"CSV 报告: {csv_path}")

    print(f"JSON 报告: {json_path}")
    print(f"Markdown 报告: {md_path}")
    print(f"状态: {report['status']}")
    print(f"ERROR: {report['summary']['errors']}, WARN: {report['summary']['warns']}, INFO: {report['summary']['infos']}")


if __name__ == "__main__":
    main()
